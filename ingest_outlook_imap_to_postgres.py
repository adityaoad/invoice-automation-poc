#!/usr/bin/env python3
"""
Outlook (IMAP) → OCR (Tesseract) → GPT extraction → PostgreSQL (normalized schema: invoice_ai.*)

Prereqs (venv active):
  brew install poppler tesseract
  pip install imapclient python-dotenv pdf2image pytesseract openai psycopg2-binary

.env example:
  OUTLOOK_EMAIL=your_name@outlook.com
  OUTLOOK_APP_PASSWORD=your_app_password
  IMAP_HOST=outlook.office365.com
  IMAP_PORT=993
  ALLOWED_SENDERS=verizon.com,att.com

  OPENAI_API_KEY=YOUR_OPENAI_KEY

  PG_DB=invoice_db
  PG_USER=postgres
  PG_PASSWORD=your_password
  PG_HOST=localhost
  PG_PORT=5432
"""

import os
import io
import json
import email
import hashlib
import pathlib
import traceback
import email.utils
from datetime import datetime, timezone


from pathlib import Path
from dotenv import load_dotenv

env_path = Path(__file__).resolve().parent / ".env"
print("Looking for .env at:", env_path)
load_dotenv(dotenv_path=env_path)

import os

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")  # always load .env next to this script

# Debug: confirm env vars loaded
print("ENV check OUTLOOK_EMAIL:", os.getenv("OUTLOOK_EMAIL"))
print("ENV check OPENAI_API_KEY present:", bool(os.getenv("OPENAI_API_KEY")))

from imapclient import IMAPClient
import psycopg2

from pdf2image import convert_from_path
import pytesseract
from openai import OpenAI

import re
from dateutil import parser as dateparser

def as_date(s):
    if not s or not str(s).strip():
        return None
    try:
        return dateparser.parse(str(s)).date()
    except Exception:
        return None

def as_decimal(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace(",", "")
    m = re.search(r"([-+]?\d+(\.\d+)?)", s)
    return float(m.group(1)) if m else None

def derive_currency(amount_str, explicit_currency):
    if explicit_currency and explicit_currency.strip():
        return explicit_currency.strip().upper()
    if not amount_str:
        return None
    txt = str(amount_str)
    if "$" in txt: return "USD"
    if "€" in txt: return "EUR"
    if "£" in txt: return "GBP"
    if "₹" in txt: return "INR"
    if "C$" in txt: return "CAD"
    if "A$" in txt: return "AUD"
    # fallback if someone passes a code like usd/cad
    c = txt.strip().upper()
    return c if len(c) == 3 else None


# --------------------------- Config / env ---------------------------

load_dotenv()  # load .env in current working directory

# Outlook IMAP
IMAP_HOST    = os.getenv("IMAP_HOST", "outlook.office365.com")
IMAP_PORT    = int(os.getenv("IMAP_PORT", "993"))
GMAIL_EMAIL = os.getenv("GMAIL_EMAIL")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")

# OpenAI
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Postgres
PG_DB       = os.getenv("PG_DB", "invoice_db")
PG_USER     = os.getenv("PG_USER", "postgres")
PG_PASSWORD = os.getenv("PG_PASSWORD", "")
PG_HOST     = os.getenv("PG_HOST", "localhost")
PG_PORT     = os.getenv("PG_PORT", "5432")

# Optional: restrict by sender domain(s)
ALLOWED_SENDERS = [d.strip().lower() for d in os.getenv("ALLOWED_SENDERS","").split(",") if d.strip()]

# Local download folder for PDFs
DOWNLOAD_DIR = pathlib.Path("inbox_downloads"); DOWNLOAD_DIR.mkdir(exist_ok=True)
PROCESSED_FOLDER = "Processed"  # IMAP folder to move processed emails

# Validate critical env
def require(name, value):
    if not value:
        raise ValueError(f"❌ Missing required env var: {name}. Check your .env.")
    return value

require("GMAIL_EMAIL", GMAIL_EMAIL)
require("GMAIL_APP_PASSWORD", GMAIL_APP_PASSWORD)
require("OPENAI_API_KEY", OPENAI_API_KEY)

print("✅ ENV: OPENAI_API_KEY loaded:", bool(OPENAI_API_KEY))
print("✅ ENV: GMAIL_EMAIL:", GMAIL_EMAIL)
print("✅ ENV: IMAP host/port:", IMAP_HOST, IMAP_PORT)

# Initialize OpenAI client
client = OpenAI(api_key=OPENAI_API_KEY)

# --------------------------- Postgres helpers ---------------------------

conn = psycopg2.connect(
    dbname=PG_DB, user=PG_USER, password=PG_PASSWORD, host=PG_HOST, port=PG_PORT
)
conn.autocommit = True
cur = conn.cursor()

def get_or_create(table, unique_key, data_dict):
    """Upsert-like helper that returns the row id for vendors/accounts/POs."""
    cols = list(data_dict.keys())
    vals = [data_dict[c] for c in cols]

    # Try to find existing by unique_key
    sel_sql = f"SELECT id FROM invoice_ai.{table} WHERE {unique_key} = %s LIMIT 1"
    cur.execute(sel_sql, (data_dict[unique_key],))
    row = cur.fetchone()
    if row:
        return row[0]

    # Insert and return id
    placeholders = ",".join(["%s"]*len(cols))
    colnames = ",".join(cols)
    ins_sql = f"INSERT INTO invoice_ai.{table} ({colnames}) VALUES ({placeholders}) RETURNING id"
    cur.execute(ins_sql, vals)
    return cur.fetchone()[0]

def insert_invoice_email_pipeline(file_name, out) -> int:
    """Insert into invoice_ai.email_pipeline_invoices and return invoice_id."""
    vendor_id = get_or_create("vendors", "name", {
        "name": out.get("Vendor Name",""),
        "address": out.get("Vendor Address","")
    })
    account_id = get_or_create("accounts", "number", {
        "number":  out.get("Account Number",""),
        "name":    out.get("Account Name",""),
        "manager": out.get("Account Manager","")
    })
    po_id = get_or_create("purchase_orders", "po_number", {
        "po_number":      out.get("PO Number",""),
        "billing_period": out.get("Billing Period",""),
        "tax_code":       out.get("Tax Code",""),
        "tax_amount":     out.get("Tax Amount","")
    })

    # ⇩ sanitize/normalize fields
    invoice_date = as_date(out.get("Invoice Date"))
    due_date     = as_date(out.get("Due Date"))
    total_amount = as_decimal(out.get("Total Amount"))
    currency     = derive_currency(out.get("Total Amount"), out.get("Currency"))

    sql = """
    INSERT INTO invoice_ai.email_pipeline_invoices (
        file, invoice_number, invoice_date, due_date, currency, total_amount,
        vendor_id, account_id, po_id
    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    ON CONFLICT (file) DO NOTHING
    RETURNING id
    """
    data = (
        file_name,
        out.get("Invoice Number"),
        invoice_date,            # None if blank → NULL (OK)
        due_date,                # None if blank → NULL (OK)
        currency,                # e.g., USD derived from "$"
        total_amount,            # numeric (float) or NULL
        vendor_id, account_id, po_id
    )
    cur.execute(sql, data)
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("SELECT id FROM invoice_ai.email_pipeline_invoices WHERE file=%s", (file_name,))
    return cur.fetchone()[0]



def insert_email_invoice(invoice_id: int, file_name: str, msg):
    """Upsert email metadata and link to email_pipeline_invoices(invoice_id)."""
    subject = msg.get("Subject", "")
    sender  = msg.get("From", "")
    mid     = msg.get("Message-ID") or msg.get("Message-Id") or ""
    # parse Date → timestamp (safe)
    try:
        dt = email.utils.parsedate_to_datetime(msg.get("Date"))
    except Exception:
        dt = None

    cur.execute("""
        INSERT INTO invoice_ai.email_invoices
            (message_id, subject, sender, received_at, invoice_id)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (message_id) DO UPDATE
           SET invoice_id = EXCLUDED.invoice_id,
               subject    = EXCLUDED.subject,
               sender     = EXCLUDED.sender,
               received_at= COALESCE(email_invoices.received_at, EXCLUDED.received_at)
    """, (mid, subject, sender, dt, invoice_id))


# --------------------------- OCR ---------------------------

def extract_text_from_pdf(pdf_path: str) -> str:
    """Convert all pages to images, OCR with Tesseract, return concatenated text."""
    images = convert_from_path(pdf_path)
    text = ""
    for img in images:
        text += pytesseract.image_to_string(img)
    return text

# --------------------------- GPT ---------------------------

def extract_fields_with_gpt(text: str) -> dict:
    """
    Use GPT to extract a rich set of invoice fields.
    Returns strict JSON with predictable keys and normalized formats.
    """
    from openai import OpenAI
    import json, re
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

    system = (
        "You are an information extraction service for invoices. "
        "Return STRICT JSON only (no prose). Use null for unknowns."
    )

    user = f"""
Extract the following fields from the invoice text.
Return a SINGLE JSON object with EXACT keys and the following constraints:

Keys (exact spelling):
- "Invoice Number": string
- "Invoice Date": string (YYYY-MM-DD if possible)
- "Due Date": string (YYYY-MM-DD or null)
- "Vendor Name": string or null
- "Vendor Address": string or null
- "PO Number": string or null
- "Billing Period": string or null
- "Account Number": string or null
- "Account Name": string or null
- "Account Manager": string or null
- "Tax Code": string or null
- "Subtotal": string or null                # numeric string, e.g., "680.00"
- "Tax Amount": string or null              # numeric string, e.g., "34.00"
- "Currency": string or null                # ISO code if obvious (USD, CAD, GBP, EUR, INR)
- "Total Amount": string                    # numeric string, required
- "Line Description": string or null        # short human label for this invoice (vendor + invno)

Normalization rules:
- Dates: prefer ISO YYYY-MM-DD if you can infer; else keep as seen.
- Amounts: output ONLY digits and a decimal point (strip currency symbols and commas).
- Currency: output a code like USD/CAD/INR/GBP/EUR if present or clearly implied; else null.
- If a field truly does not appear, set it to null.

Example output:
{{
  "Invoice Number": "INV-101905",
  "Invoice Date": "2025-08-05",
  "Due Date": "2025-08-24",
  "Vendor Name": "Scott Inc",
  "Vendor Address": "123 Example St, Toronto, ON",
  "PO Number": "PO-77831",
  "Billing Period": "2025-07",
  "Account Number": "5850133469",
  "Account Name": "Innovate Wireless Partnerships",
  "Account Manager": "Cody Burke",
  "Tax Code": "VAT-20%",
  "Subtotal": "680.00",
  "Tax Amount": "34.00",
  "Currency": "CAD",
  "Total Amount": "714.00",
  "Line Description": "Scott Inc — INV-101905"
}}

INVOICE TEXT:
\"\"\"{text}\"\"\"
"""

    resp = client.chat.completions.create(
        model="gpt-4o-mini",  # or "gpt-4.1-mini" if you have access
        temperature=0,
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    )
    raw = resp.choices[0].message.content.strip()

    # Be tolerant to ```json fences
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw, flags=re.IGNORECASE | re.DOTALL)

    try:
        data = json.loads(raw)
    except Exception:
        # last-ditch: try to locate the first {...}
        m = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        data = json.loads(m.group(0)) if m else {}

    # post-normalize numeric strings (strip $ and commas just in case)
    def clean_amt(v):
        if v is None: return None
        s = str(v).replace("$","").replace(",","").strip()
        # allow bare numbers only
        return s

    for k in ["Subtotal", "Tax Amount", "Total Amount"]:
        if k in data:
            data[k] = clean_amt(data[k])

    # ensure required keys exist even if null
    required = ["Invoice Number","Invoice Date","Due Date","Vendor Name","Vendor Address",
                "PO Number","Billing Period","Account Number","Account Name","Account Manager",
                "Tax Code","Subtotal","Tax Amount","Currency","Total Amount","Line Description"]
    for k in required:
        data.setdefault(k, None)

    # sensible fallback for Line Description
    if not data.get("Line Description"):
        vn = (data.get("Vendor Name") or "Vendor").strip()
        ino = (data.get("Invoice Number") or "").strip()
        data["Line Description"] = f"{vn} — {ino}".strip(" —")

    return data


# --------------------------- IMAP helpers ---------------------------

def ensure_folder(server: IMAPClient, folder: str):
    try:
        server.create_folder(folder)
    except Exception:
        pass  # already exists

def sender_allowed(from_header: str) -> bool:
    if not ALLOWED_SENDERS:
        return True
    try:
        addr = email.utils.parseaddr(from_header)[1].lower()
        return any(addr.endswith("@"+d) for d in ALLOWED_SENDERS)
    except:
        return True

def save_pdf_attachments(msg):
    """Save PDF attachments to DOWNLOAD_DIR, return list of saved file paths."""
    saved = []
    if msg.is_multipart():
        for part in msg.walk():
            cdisp = part.get("Content-Disposition", "")
            if "attachment" in cdisp:
                fname = part.get_filename()
                if not fname:
                    continue
                if not fname.lower().endswith(".pdf"):
                    continue
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                path = DOWNLOAD_DIR / fname
                if path.exists():
                    ts = datetime.now().strftime("%Y%m%d%H%M%S")
                    path = DOWNLOAD_DIR / f"{ts}_{fname}"
                with open(path, "wb") as f:
                    f.write(payload)
                saved.append(str(path))
    return saved

from openpyxl import Workbook, load_workbook

AP_XLSX_COLS = [
    "Invoice number", "Supplier number", "Supplier site", "Description",
    "Pay group", "Type", "Amount", "Line description", "Entity",
    "Region", "Function", "Expense account", "Product", "Project",
    "Intercompany", "Future use", "N/A"
]

AP_DEFAULTS = {
    "Supplier site": "HQ",
    "Description": "AP Load",
    "Pay group": "STANDARD",
    "Entity": "US",
    "Region": "NA",
    "Expense account": "6100",
    "Product": "GEN",
    "Project": "None",
    "Intercompany": "No",
    "Future use": "",
    "N/A": ""
}

def ensure_ap_workbook(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "AP"
        ws.append(AP_XLSX_COLS)
        wb.save(path)

from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime
import random

from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime
import os, random

from datetime import datetime

def normalize_date(s: str) -> str:
    if not s: 
        return ""
    s = s.strip()
    fmts = [
        "%Y-%m-%d", "%Y/%m/%d",                # 2025-09-10
        "%m-%d-%Y", "%m/%d/%Y",                # 09-10-2025
        "%d-%m-%Y", "%d/%m/%Y",                # 10-09-2025
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"
    ]
    for f in fmts:
        try:
            dt = datetime.strptime(s, f)
            return dt.strftime("%m/%d/%y")     # Excel-friendly format
        except Exception:
            pass
    return s  # if it’s some other readable format, keep as-is


XLSX_PATH = "/Users/adityasmacbookair/Documents/Invoice Automation Project/test book.xlsx"

def append_ap_rows_to_excel(xlsx_path, fields):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    pos = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def to_dec(x):
        s = ("" if x is None else str(x)).replace(",", "").replace("$", "").strip()
        try:
            return Decimal(s) if s else Decimal("0")
        except Exception:
            return Decimal("0")

    inv_no    = (fields.get("Invoice Number") or "").strip()
    supp_no   = (fields.get("Supplier Number") or f"333{random.randint(100,999)}").strip()
    line_desc = (fields.get("Line Description") or f"{fields.get('Vendor Name','Vendor')} — {inv_no}").strip()
    func_val  = (fields.get("Function") or "9600").strip()
    total     = to_dec(fields.get("Total Amount"))
    tax       = to_dec(fields.get("Tax Amount"))
    item      = max(total - tax, Decimal("0"))
    inv_date_norm = normalize_date(fields.get("Invoice Date"))
    fixed = {
        "invoice date": inv_date_norm or "",       # <-- use extracted invoice date,
        "supplier site": "JPY59",
        "description": "",            # D stays empty only if header is "description"; D is blank in your file anyway
        "pay group": "AP3828",
        "today's date": datetime.now().strftime("%m/%d/%y"),
        "entity": "2827",
        "region": "510",
        "expense account": "725",
        "product": "1100",
        "project": "0",
        "intercompany": "0",
        "future use": "0",
        "n/a": "NO",
    }

    def build_row(row_type, amount):
        row = [""] * len(headers)
        for h_raw, idx in pos.items():
            # keep any blank/None header columns empty (covers D, I, etc.)
            if h_raw in ("", "none"):
                row[idx] = ""
                continue

            # dynamic fields
            if h_raw == "invoice number":           row[idx] = inv_no
            elif h_raw == "supplier number":        row[idx] = supp_no
            elif h_raw in ("line description","line desctiption"):
                                                    row[idx] = line_desc
            elif h_raw == "function":               row[idx] = func_val
            elif h_raw == "type":                   row[idx] = row_type
            elif h_raw == "amount":                 row[idx] = float(amount)
            # fixed named headers
            elif h_raw in fixed:                    row[idx] = fixed[h_raw]
            else:                                   row[idx] = ""     # anything unknown stays empty
        return row

    ws.append(build_row("ITEM", item))
    ws.append(build_row("TAX",  tax))
    wb.save(xlsx_path)
    print("✅ wrote 2 rows to:", xlsx_path)



if not os.path.exists(XLSX_PATH):
    raise SystemExit(f"File not found: {XLSX_PATH}")

# --------------------------- Main ---------------------------

def main():
    print("Connecting to IMAP…")
    with IMAPClient(IMAP_HOST, port=IMAP_PORT, use_uid=True, ssl=True) as server:
        server.login(GMAIL_EMAIL, GMAIL_APP_PASSWORD)
        label = os.getenv("GMAIL_LABEL", "INBOX")
        server.select_folder(label)
        print("Using label:", label)
        print("Unread count:", len(server.search("UNSEEN")))
        ensure_folder(server, PROCESSED_FOLDER)

        # Fetch unread messages
        uids = server.search("UNSEEN")
        if not uids:
            print("No unread messages.")
            return

        for uid in uids:
            try:
                raw = server.fetch(uid, ["RFC822"])[uid][b"RFC822"]
                msg = email.message_from_bytes(raw)

                from_header = msg.get("From","")
                subject = msg.get("Subject","")
                if not sender_allowed(from_header):
                    print(f"Skipping sender: {from_header}")
                    server.add_flags(uid, [b"\\Seen"])
                    server.copy(uid, PROCESSED_FOLDER)
                    server.delete_messages(uid)
                    continue

                pdfs = save_pdf_attachments(msg)
                if not pdfs:
                    print(f"No PDF attachments in: {subject}")
                    server.add_flags(uid, [b"\\Seen"])
                    server.copy(uid, PROCESSED_FOLDER)
                    server.delete_messages(uid)
                    continue

                for pdf_path in pdfs:
                    file_name = os.path.basename(pdf_path)
                    print(f"Processing {file_name}")
                    text   = extract_text_from_pdf(pdf_path)
                    fields = extract_fields_with_gpt(text)
                    import json
                    print("[FIELDS]", json.dumps(fields, ensure_ascii=False))
                    inv_no = (fields.get("Invoice Number") or "").strip()
                    print("[KEYS] inv:", inv_no, "total:", fields.get("Total Amount"), "tax:", fields.get("Tax Amount"))


                    # 1) Save to Postgres
                    invoice_id = insert_invoice_email_pipeline(file_name, fields)
                    insert_email_invoice(invoice_id, file_name, msg)

                    # 2) Save to Excel as well
                    xlsx_path = os.getenv("SHAREPOINT_XLSX")
                    append_ap_rows_to_excel(xlsx_path, fields)
                    print("[EXCEL] appended for", inv_no, "→", xlsx_path)



                # Mark handled: read + move to Processed + delete original
                server.add_flags(uid, [b"\\Seen"])
                server.copy(uid, PROCESSED_FOLDER)
                server.delete_messages(uid)

                print("✅ Message processed and moved.")

            except Exception as e:
                print("❌ Error handling message:", e)
                traceback.print_exc()

        server.expunge()
        print("Done.")

if __name__ == "__main__":
    main()