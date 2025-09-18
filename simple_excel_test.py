from openpyxl import Workbook, load_workbook
import os

xlsx_path = "simple_test.xlsx"

# if file doesn't exist, create with headers
if not os.path.exists(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws.append(["Invoice Number", "Supplier Number", "Type", "Amount"])
    wb.save(xlsx_path)
    print("Created new file:", xlsx_path)

# reopen and append a test row
wb = load_workbook(xlsx_path)
ws = wb.active
ws.append(["INV-001", "SUP-123", "Item", 4500])
ws.append(["INV-001", "SUP-123", "Tax", 500])
wb.save(xlsx_path)
print("✅ Added two rows to", xlsx_path)
