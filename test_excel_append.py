from openpyxl import load_workbook
from decimal import Decimal
import random


def append_ap_rows_to_excel(xlsx_path, fields):
    print("[xlsx] target:", xlsx_path)

    """
    Appends 2 rows per invoice to the ACTIVE sheet in xlsx_path.
    Uses the header row (row 1) to map columns by name (case-insensitive).
    Dynamic: 'invoice number', 'supplier number', 'line desctiption', 'function'
    Real split for: 'type' (Item/Tax) and 'amount'
    Everything else = 'Fixed Value'.
    """
    # --- pull dynamic values from extracted fields (with safe defaults)
    inv_no   = (fields.get("Invoice Number") or "").strip()
    vendor   = (fields.get("Vendor Name") or fields.get("Vendor") or "Unknown Vendor").strip()
    inv_date = (fields.get("Invoice Date") or "").strip()
    supp_no  = (fields.get("Supplier Number") or f"SUP-{random.randint(100000,999999)}").strip()
    line_desc= (fields.get("Line Description") or f"{vendor} — {inv_no} — {inv_date}").strip()
    func_val = (fields.get("Function") or "IT").strip()

    # amounts
    def to_dec(x):
        s = ("" if x is None else str(x)).replace(",", "").replace("$", "").strip()
        try:
            return Decimal(s) if s else Decimal("0")
        except Exception:
            return Decimal("0")
    total = to_dec(fields.get("Total Amount"))
    tax   = to_dec(fields.get("Tax Amount"))
    item  = max(total - tax, Decimal("0"))

    # --- open workbook & read headers
    wb = load_workbook(xlsx_path)
    ws = wb.active  # use current active sheet
    headers = [ (c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1)) ]
    # map lowercased header -> index
    idx = { str(h).strip().lower(): i for i, h in enumerate(headers) }

    # helper to build a row initialized to 'Fixed Value' everywhere
    def base_row():
        return ["Fixed Value" for _ in headers]

    # write one row given row_type and amount
    def build_row(row_type: str, amount: Decimal):
        row = base_row()

        # set the four dynamic fields (A,B,H,K by header name)
        if "invoice number" in idx:      row[idx["invoice number"]] = inv_no
        if "supplier number" in idx:     row[idx["supplier number"]] = supp_no
        if "line desctiption" in idx:    row[idx["line desctiption"]] = line_desc  # keep typo as provided
        if "function" in idx:            row[idx["function"]] = func_val

        # real split columns
        if "type" in idx:                row[idx["type"]] = row_type
        if "amount" in idx:              row[idx["amount"]] = float(amount)

        return row

    ws.append(build_row("Item", item))
    ws.append(build_row("Tax",  tax))
    wb.save(xlsx_path)
    print("[xlsx] appended 2 rows and saved")

