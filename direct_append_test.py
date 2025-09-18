from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime
import os, random

XLSX_PATH = "/Users/adityasmacbookair/Documents/Invoice Automation Project/test book.xlsx"

def append_ap_rows_to_excel(xlsx_path, fields):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    pos = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def to_dec(x):
        s = ("" if x is None else str(x)).replace(",", "").replace("$", "").strip()
        try:
            return Decimal(s) if s else Decimal("0")
        except Exception:
            return Decimal("0")

    inv_no    = (fields.get("Invoice Number") or "").strip()
    supp_no   = (fields.get("Supplier Number") or f"333{random.randint(100,999)}").strip()
    line_desc = (fields.get("Line Description") or f"{fields.get('Vendor Name','Vendor')} — {inv_no}").strip()
    func_val  = (fields.get("Function") or "9600").strip()
    total     = to_dec(fields.get("Total Amount"))
    tax       = to_dec(fields.get("Tax Amount"))
    item      = max(total - tax, Decimal("0"))

    fixed = {
        "invoice date": datetime.now().strftime("%m/%d/%y"),
        "supplier site": "JPY59",
        "description": "",            # D stays empty only if header is "description"; D is blank in your file anyway
        "pay group": "AP3828",
        "today's date": datetime.now().strftime("%m/%d/%y"),
        "entity": "2827",
        "region": "510",
        "expense account": "725",
        "product": "1100",
        "project": "0",
        "intercompany": "0",
        "future use": "0",
        "n/a": "NO",
    }

    def build_row(row_type, amount):
        row = [""] * len(headers)
        for h_raw, idx in pos.items():
            # keep any blank/None header columns empty (covers D, I, etc.)
            if h_raw in ("", "none"):
                row[idx] = ""
                continue

            # dynamic fields
            if h_raw == "invoice number":           row[idx] = inv_no
            elif h_raw == "supplier number":        row[idx] = supp_no
            elif h_raw in ("line description"):
                                                    row[idx] = line_desc
            elif h_raw == "function":               row[idx] = func_val
            elif h_raw == "type":                   row[idx] = row_type
            elif h_raw == "amount":                 row[idx] = float(amount)
            # fixed named headers
            elif h_raw in fixed:                    row[idx] = fixed[h_raw]
            else:                                   row[idx] = ""     # anything unknown stays empty
        return row

    ws.append(build_row("ITEM", item))
    ws.append(build_row("TAX",  tax))
    wb.save(xlsx_path)
    print("✅ wrote 2 rows to:", xlsx_path)

fields = {
    "Invoice Number": "TEST-LOCAL-003",
    "Vendor Name": "Test Vendor",
    "Invoice Date": "2025-07-31",
    "Total Amount": "$5,000",
    "Tax Amount": "$500",
}

if not os.path.exists(XLSX_PATH):
    raise SystemExit(f"File not found: {XLSX_PATH}")

append_ap_rows_to_excel(XLSX_PATH, fields)
