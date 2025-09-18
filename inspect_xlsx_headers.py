import os
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
p = os.getenv("SHAREPOINT_XLSX")
print("Path:", p)

wb = load_workbook(p)
ws = wb.active
print("Active sheet:", ws.title)

row1 = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("Headers (raw):", row1)
print("Headers (debug repr):", [repr(x) for x in row1])
print("Header -> index:", {str(h).strip().lower(): i for i, h in enumerate(row1)})
