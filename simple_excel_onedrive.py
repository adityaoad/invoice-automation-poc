import os
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()
xlsx_path = os.getenv("SHAREPOINT_XLSX")  # from your .env

print("Target path:", xlsx_path)

# if file doesn't exist, create with headers
if not os.path.exists(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws.append(["Invoice Number", "Supplier Number", "Type", "Amount"])
    wb.save(xlsx_path)
    print("Created new file at OneDrive location")

# reopen and append test rows
wb = load_workbook(xlsx_path)
ws = wb.active
ws.append(["INV-OD-001", "SUP-999", "Item", 1234])
ws.append(["INV-OD-001", "SUP-999", "Tax", 123])
wb.save(xlsx_path)
print("✅ Added two rows to OneDrive file")
